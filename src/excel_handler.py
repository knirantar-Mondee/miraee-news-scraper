import os
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from config.settings import (
    RSS_FEEDS_FILE, COMPETITORS_FILE, SEARCH_QUERIES_FILE, DATABASE_FILE, OUTPUT_DIR
)
from src.utils import logger

def ensure_input_templates():
    """Ensure that the input files exist with default templates if not present."""
    os.makedirs(os.path.dirname(RSS_FEEDS_FILE), exist_ok=True)
    
    # 1. RSS Feeds template
    if not os.path.exists(RSS_FEEDS_FILE):
        df_rss = pd.DataFrame([
            {"Feed_ID": 1, "Feed_Name": "Skift", "RSS_URL": "https://skift.com/feed", "Category": "Travel", "Active": "Yes"},
            {"Feed_ID": 2, "Feed_Name": "PhocusWire", "RSS_URL": "https://www.phocuswire.com/rss", "Category": "Travel", "Active": "Yes"},
            {"Feed_ID": 3, "Feed_Name": "Travel Weekly", "RSS_URL": "https://www.travelweekly.com/rss", "Category": "Travel", "Active": "Yes"}
        ])
        with pd.ExcelWriter(RSS_FEEDS_FILE, engine='openpyxl') as writer:
            df_rss.to_excel(writer, sheet_name="RSS_Feeds", index=False)
        logger.info(f"Created template: {RSS_FEEDS_FILE}")
        
    # 2. Competitors template
    if not os.path.exists(COMPETITORS_FILE):
        df_comp = pd.DataFrame([
            {"Business": "B2E", "Competitor": "Navan", "Base": "Miraee"},
            {"Business": "B2E", "Competitor": "TravelPerk", "Base": "Miraee"},
            {"Business": "B2E", "Competitor": "SAP Concur", "Base": "Miraee"},
            {"Business": "B2E", "Competitor": "Ramp", "Base": "Miraee"},
            {"Business": "B2E", "Competitor": "Brex", "Base": "Miraee"},
            {"Business": "B2C", "Competitor": "AirBnb", "Base": "Abhee"},
            {"Business": "B2C", "Competitor": "TripGenie", "Base": "Abhee"},
            {"Business": "B2C", "Competitor": "Expedia", "Base": "Abhee"},
            {"Business": "B2C", "Competitor": "Booking.com", "Base": "Abhee"},
            {"Business": "B2B", "Competitor": "Centrav", "Base": "Mondee"},
            {"Business": "B2B", "Competitor": "Picasso Travel", "Base": "Mondee"},
            {"Business": "B2B", "Competitor": "Sky Bird Travel", "Base": "Mondee"},
            {"Business": "B2B", "Competitor": "GTT Global", "Base": "Mondee"},
            {"Business": "B2B", "Competitor": "Downtown Travel", "Base": "Mondee"},
            {"Business": "B2B", "Competitor": "Expedia TAAP", "Base": "Mondee"},
            {"Business": "B2B", "Competitor": "Booking.com for Travel Agents", "Base": "Mondee"},
            {"Business": "B2B", "Competitor": "Priceline Partner Network", "Base": "Mondee"}
        ])
        with pd.ExcelWriter(COMPETITORS_FILE, engine='openpyxl') as writer:
            df_comp.to_excel(writer, sheet_name="Competitors", index=False)
        logger.info(f"Created template: {COMPETITORS_FILE}")
        
    # 3. Search Queries template
    if not os.path.exists(SEARCH_QUERIES_FILE):
        df_queries = pd.DataFrame([
            {"Competitor": "Mondee", "Query_Type": "Brand", "Search_Query": "Mondee"},
            {"Competitor": "Mondee", "Query_Type": "Product", "Search_Query": "Abhi"},
            {"Competitor": "Mondee", "Query_Type": "Product", "Search_Query": "Rocketrip"},
            {"Competitor": "Mondee", "Query_Type": "Product", "Search_Query": "TripPlanet"},
            {"Competitor": "Mondee", "Query_Type": "Executive", "Search_Query": "Prasad Gundumogula"},
            {"Competitor": "Mondee", "Query_Type": "Executive", "Search_Query": "Oren Zeev"},
            {"Competitor": "Mondee", "Query_Type": "Technology", "Search_Query": "AI Travel Marketplace"},
            
            {"Competitor": "Navan", "Query_Type": "Brand", "Search_Query": "Navan"},
            {"Competitor": "Navan", "Query_Type": "Brand", "Search_Query": "TripActions"},
            {"Competitor": "Navan", "Query_Type": "Product", "Search_Query": "Navan Connect"},
            
            {"Competitor": "TravelPerk", "Query_Type": "Brand", "Search_Query": "TravelPerk"},
            {"Competitor": "TravelPerk", "Query_Type": "Product", "Search_Query": "GreenPerk"}
        ])
        with pd.ExcelWriter(SEARCH_QUERIES_FILE, engine='openpyxl') as writer:
            df_queries.to_excel(writer, sheet_name="Search_Queries", index=False)
        logger.info(f"Created template: {SEARCH_QUERIES_FILE}")

def load_rss_feeds():
    """Load active RSS feeds from Excel."""
    ensure_input_templates()
    try:
        df = pd.read_excel(RSS_FEEDS_FILE, sheet_name="RSS_Feeds")
        # Filter for active feeds only
        active_df = df[df["Active"].astype(str).str.strip().str.lower() == "yes"]
        logger.info(f"Loaded {len(active_df)} active RSS feeds out of {len(df)} total.")
        return active_df.to_dict('records')
    except Exception as e:
        logger.error(f"Error loading RSS feeds: {e}")
        return []

def load_competitors():
    """Load competitors list from Excel."""
    ensure_input_templates()
    try:
        df = pd.read_excel(COMPETITORS_FILE, sheet_name="Competitors")
        competitors = df["Competitor"].dropna().astype(str).str.strip().tolist()
        logger.info(f"Loaded {len(competitors)} competitors.")
        return competitors
    except Exception as e:
        logger.error(f"Error loading competitors: {e}")
        return []

def load_competitor_brands():
    """Load competitor to target brand (Base) mapping from Excel."""
    ensure_input_templates()
    try:
        df = pd.read_excel(COMPETITORS_FILE, sheet_name="Competitors")
        mapping = {}
        col_to_use = "Base" if "Base" in df.columns else ("Business" if "Business" in df.columns else None)
        for _, row in df.iterrows():
            competitor = str(row["Competitor"]).strip()
            if col_to_use:
                base = str(row[col_to_use]).strip()
            else:
                base = "Miraee"
            if competitor and competitor.lower() != 'nan':
                mapping[competitor] = base
        logger.info(f"Loaded brand mapping for {len(mapping)} competitors.")
        return mapping
    except Exception as e:
        logger.error(f"Error loading competitor brands: {e}")
        return {}

def load_search_queries():
    """Load search queries from Excel and build mapping: {Competitor: [(Query, Type), ...]}."""
    ensure_input_templates()
    try:
        df = pd.read_excel(SEARCH_QUERIES_FILE, sheet_name="Search_Queries")
        queries_dict = {}
        for _, row in df.iterrows():
            competitor = str(row["Competitor"]).strip()
            query = str(row["Search_Query"]).strip()
            q_type = str(row["Query_Type"]).strip() if "Query_Type" in row else "General"
            
            if competitor and query and query.lower() != 'nan':
                if competitor not in queries_dict:
                    queries_dict[competitor] = []
                queries_dict[competitor].append((query, q_type))
                
        # Ensure all listed competitors have at least one query
        competitors = load_competitors()
        for competitor in competitors:
            if competitor not in queries_dict:
                queries_dict[competitor] = [(competitor, "Brand")]
                logger.info(f"No custom queries found for competitor '{competitor}'. Auto-generated default query: '{competitor}' (Brand).")
                
        logger.info(f"Loaded query mappings for {len(queries_dict)} competitors.")
        return queries_dict
    except Exception as e:
        logger.error(f"Error loading search queries: {e}")
        return {}

def read_existing_urls():
    """Read already scraped article URLs from all brand output databases and legacy database for deduplication."""
    urls = set()
    import glob
    db_pattern = os.path.join(OUTPUT_DIR, "raw_news_database_*.xlsx")
    db_files = glob.glob(db_pattern)
    legacy_file = os.path.join(OUTPUT_DIR, "raw_news_database.xlsx")
    if os.path.exists(legacy_file):
        db_files.append(legacy_file)
        
    for db_file in db_files:
        try:
            df = pd.read_excel(db_file, sheet_name="Raw_News", usecols=["Article_URL"])
            urls.update(df["Article_URL"].dropna().astype(str).str.strip().tolist())
        except Exception as e:
            logger.warning(f"Could not read existing database URLs from {db_file}: {e}")
    logger.info(f"Found {len(urls)} existing URLs across all databases.")
    return urls

def append_news_articles_to_file(articles, file_path, brand):
    """Append articles to a specific Excel database file."""
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    headers = [
        "Article_ID", "RSS_Source", "RSS_URL", "Article_URL", "Published_Date", 
        "Competitor", "Matched_Query", "Title", "Author", "News_Body", "Scrape_Date",
        "Topic", "Sentiment", "Threat_Level", "Competitor_Action", "Strategic_Implication", "Executive_Summary",
        "Target_Brand"
    ]
    
    file_exists = os.path.exists(file_path)
    
    try:
        if not file_exists:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Raw_News"
            ws.append(headers)
            wb.save(file_path)
            wb.close()
            logger.info(f"Initialized new database file for {brand}: {file_path}")
            
        # Open workbook and append rows
        wb = openpyxl.load_workbook(file_path)
        ws = wb["Raw_News"]
        
        # Self-healing header check: ensure old databases are updated with new columns
        for idx, h in enumerate(headers, 1):
            if ws.cell(row=1, column=idx).value != h:
                ws.cell(row=1, column=idx).value = h
        
        # Determine the starting Article_ID
        max_row = ws.max_row
        start_id = 1
        if max_row > 1:
            try:
                # Retrieve last ID from column A
                last_id_val = ws.cell(row=max_row, column=1).value
                start_id = int(last_id_val) + 1
            except (ValueError, TypeError):
                start_id = max_row # fallback
                
        for article in articles:
            row_data = [
                start_id,
                article.get("RSS_Source", ""),
                article.get("RSS_URL", ""),
                article.get("Article_URL", ""),
                article.get("Published_Date", ""),
                article.get("Competitor", ""),
                article.get("Matched_Query", ""),
                article.get("Title", ""),
                article.get("Author", ""),
                article.get("News_Body", ""),
                article.get("Scrape_Date", ""),
                article.get("Topic", ""),
                article.get("Sentiment", ""),
                article.get("Threat_Level", ""),
                article.get("Competitor_Action", ""),
                article.get("Strategic_Implication", ""),
                article.get("Executive_Summary", ""),
                article.get("Target_Brand", "")
            ]
            ws.append(row_data)
            start_id += 1
            
        wb.save(file_path)
        wb.close()
        logger.info(f"Successfully appended {len(articles)} articles to database: {file_path}")
    except Exception as e:
        logger.error(f"Error appending articles to Excel database {file_path}: {e}")
        # Try to save to a backup CSV file in case Excel writing completely locks
        try:
            backup_csv = file_path.replace(".xlsx", "_backup.csv")
            df_new = pd.DataFrame(articles)
            df_new.to_csv(backup_csv, mode='a', header=not os.path.exists(backup_csv), index=False)
            logger.info(f"Saved backup articles data to CSV: {backup_csv}")
        except Exception as csv_err:
            logger.critical(f"Backup writing failed: {csv_err}")

def append_news_articles(articles):
    """Group articles by brand and append to separate database Excel files."""
    if not articles:
        return
        
    # Group articles by brand
    brand_groups = {}
    for article in articles:
        target_brand_str = article.get("Target_Brand", "Miraee")
        # Split by comma for cases where it maps to multiple brands
        brands = [b.strip() for b in target_brand_str.split(",") if b.strip()]
        if not brands:
            brands = ["Miraee"] # Default fallback
            
        for brand in brands:
            if brand not in brand_groups:
                brand_groups[brand] = []
            brand_groups[brand].append(article)
            
    # Save each group to its brand-specific Excel database
    for brand, group in brand_groups.items():
        filename = f"raw_news_database_{brand}.xlsx"
        db_file_path = os.path.join(OUTPUT_DIR, filename)
        append_news_articles_to_file(group, db_file_path, brand)
        
        # Also sync to the News_Dashboard data folder for immediate UI refresh
        dashboard_dir = os.path.join(os.path.dirname(BASE_DIR), "News_Dashboard", "data")
        if os.path.exists(os.path.dirname(dashboard_dir)):
            os.makedirs(dashboard_dir, exist_ok=True)
            dashboard_file_path = os.path.join(dashboard_dir, filename)
            append_news_articles_to_file(group, dashboard_file_path, brand)
